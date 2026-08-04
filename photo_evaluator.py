"""
Photo Evaluator
Evaluates a jobsite photo against the company's structured QA/QC checklist.

Evaluation process:
  1. Load the Issue / Assessment Criteria list from library/QAQC_Agent_Training.xlsx
     ("Criteria" tab) — this is the checklist every photo is run against.
  2. Load training examples from the same workbook ("Training materials" tab) and attach the
     ones relevant to each issue, so Claude has calibrated examples of "Defect" vs "Good
     Practice" for that specific issue.
  3. Load relevant entries from the internal feedback library (library/QAQC_Library_Comments.xlsx
     + library/photos/) — prior confirmed findings from real evaluations, used as secondary,
     company-specific precedent (see append_to_library()).
  4. Load submittal context from /submittals — used to confirm correct materials/equipment where
     an assessment criterion calls for it.
  5. Web search is available as a fallback: for NYC codes / industry guidance an assessment
     criterion calls for, for anything the library/training doesn't cover, and to separately scan
     for additional QA/QC or safety issues not on the checklist.

Requires: pip install anthropic openpyxl
Requires: ANTHROPIC_API_KEY environment variable set to a real key from console.anthropic.com
(NOT the placeholder "sk-ant-...").
"""
import base64
import mimetypes
import os
import re
import shutil
from datetime import date, datetime
from pathlib import Path

import anthropic
import openpyxl
from openpyxl.utils import get_column_letter

LIBRARY_DIR = Path("library")
LIBRARY_PHOTOS_DIR = LIBRARY_DIR / "photos"
LIBRARY_XLSX = LIBRARY_DIR / "QAQC_Library_Comments.xlsx"      # stored feedback from past evaluations
LIBRARY_HEADERS = ["Photo Filename", "CSI Division / Category", "Comment / Finding",
                    "Severity", "Keywords / Tags", "Date Added", "Added By"]

CRITERIA_XLSX = LIBRARY_DIR / "QAQC_Agent_Training.xlsx"       # the company's Issue/Criteria checklist
SUBMITTALS_DIR = Path("submittals")

MODEL = "claude-sonnet-4-6"
MAX_LIBRARY_MATCHES = 5      # stored-feedback entries to include
MAX_LIBRARY_IMAGES = 3       # matched feedback reference photos to actually attach (cost control)
MAX_TRAINING_EXAMPLES_PER_ISSUE = 3
MAX_SUBMITTAL_CHARS = 4000

SYSTEM_PROMPT = """You are a construction Quality Assurance / Quality Control (QA/QC) inspector \
assistant for Cauldwell Wingate. You evaluate a jobsite photo against the company's structured \
QA/QC issue checklist.

You are given, in this order of use:

1. AN ISSUE CHECKLIST with assessment criteria (from the company's QA/QC Agent Training \
spreadsheet). Evaluate the photo against EVERY issue in this list, in the order given. This is \
the authoritative checklist — address every issue, even if only to say it isn't visible here.
2. TRAINING EXAMPLES for issues that have them: real prior photos annotated as "Defect" or "Good \
Practice," with severity, location, and observation notes. Use these to calibrate what counts as \
a defect vs. an acceptable condition for that issue. They are reference examples only — do not \
describe or evaluate the training example photos themselves, only the new photo you're given.
3. STORED FEEDBACK from previous evaluations (the company's internal library), where a relevant \
match was found — secondary, company-specific precedent, used alongside the checklist above.
4. SUBMITTAL CONTEXT, where provided — use it to confirm the correct materials/equipment were \
installed, wherever an assessment criterion calls for checking submittals or it's otherwise \
necessary. If a relevant submittal is on file only as an image/PDF that wasn't provided as text, \
say so and recommend a manual cross-check rather than guessing.
5. WEB SEARCH — use it to consult NYC building codes, industry guidelines (ACI, AISC, IBC, MSS, \
NEC, OSHA, etc.) and other authorities: (a) whenever an assessment criterion tells you to, (b) \
whenever you need it to evaluate a criterion the library/training doesn't fully cover, and (c) to \
separately scan the photo for other QA/QC or safety issues not on the given checklist.

For EACH issue on the checklist, in order, decide ONE of three outcomes:
- ISSUE FOUND — the condition is visible and does NOT meet the assessment criteria (a single \
issue can have multiple separate incidents in one photo — describe each one)
- UNCLEAR / FIELD TEAM TO VERIFY — the condition is relevant here but genuinely can't be \
determined from this photo alone (angle, resolution, lighting, occlusion)
- NO ISSUES FOUND — the condition either isn't present/visible in this photo, or is visible and \
fully meets the assessment criteria

Be conservative: only mark ISSUE FOUND when a defect is clearly visually evidenced. Never force \
a finding — most issues in a given photo will likely land in NO ISSUES FOUND.

Then organize your ENTIRE response into these four sections, using exactly these headers, in \
this order:

## Issues Found
One detailed entry per issue marked ISSUE FOUND: name the issue, describe what's wrong \
(multiple incidents if applicable), and cite the relevant training example, stored feedback, \
submittal, or web-sourced standard where applicable. Also include here, clearly prefixed \
"Additional (not on checklist):", any further problems you identify from broader review / web \
search that aren't on the given checklist. If nothing belongs in this section, write "None."

## Unclear / Field Team to Verify
One line per issue marked UNCLEAR — no more than one sentence each: just the issue name and what \
needs field verification. If nothing belongs in this section, write "None."

## No Issues Found
A plain list of the issue names marked NO ISSUES FOUND — names only, no explanation.

## Follow-Up Recommended
Concrete follow-up actions for anything listed as unclear above, and for anything that looks like \
work still in progress (not yet ready for a final finding) even if it was marked ISSUE FOUND or \
UNCLEAR. If nothing belongs in this section, write "None."
"""


# --------------------------------------------------------------------------------------
# Criteria / training spreadsheet loading
# --------------------------------------------------------------------------------------

def _read_sheet_as_records(ws, required_headers: list[str]) -> list[dict]:
    """Generic reader: finds the header row (searched within the first 15 rows) that contains
    all required_headers (case-insensitive), then reads every non-empty row below it into a
    dict keyed by those header names. Works regardless of which columns the sheet uses, so it
    keeps working if Tim adds/removes columns in the spreadsheet later."""
    header_row_num = None
    col_map = {}
    for row in ws.iter_rows(min_row=1, max_row=15):
        lowered = {str(c.value).strip().lower(): c.column for c in row if c.value is not None}
        if all(h.lower() in lowered for h in required_headers):
            header_row_num = row[0].row
            col_map = {h: lowered[h.lower()] for h in required_headers}
            break
    if header_row_num is None:
        return []

    records = []
    for row in ws.iter_rows(min_row=header_row_num + 1):
        row_num = row[0].row
        record = {}
        has_value = False
        for h, col_idx in col_map.items():
            val = ws.cell(row=row_num, column=col_idx).value
            record[h] = val
            if val not in (None, ""):
                has_value = True
        if has_value:
            records.append(record)
    return records


def _norm_issue(text) -> str:
    return re.sub(r"\s+", " ", str(text or "")).strip().lower()


def load_criteria() -> list[dict]:
    """Read the 'Criteria' tab: list of {'Issue': ..., 'Assessment criteria': ...}."""
    if not CRITERIA_XLSX.exists():
        return []
    wb = openpyxl.load_workbook(CRITERIA_XLSX, data_only=True)
    if "Criteria" not in wb.sheetnames:
        return []
    return _read_sheet_as_records(wb["Criteria"], ["Issue", "Assessment criteria"])


def diagnose_criteria_setup() -> str:
    """Human-readable diagnostic explaining why the checklist can't be loaded. Returns an empty
    string if everything checks out. Call this BEFORE evaluate_photo() so setup problems show up
    clearly instead of surfacing as a vague 'no checklist' note inside the AI's report."""
    resolved = CRITERIA_XLSX.resolve()
    if not CRITERIA_XLSX.exists():
        return (
            f"Can't find the checklist file. Looked for it at:\n    {resolved}\n"
            f"Make sure QAQC_Agent_Training.xlsx is saved in the library/ folder, and that you're "
            f"running main.py from the project's root folder (the one containing library/, input/, "
            f"checklists/, etc.)."
        )
    try:
        wb = openpyxl.load_workbook(CRITERIA_XLSX, data_only=True)
    except Exception as e:
        return f"Found the file at {resolved} but couldn't open it: {e}"

    if "Criteria" not in wb.sheetnames:
        return (
            f"Opened {resolved}, but it has no tab named exactly 'Criteria' "
            f"(found tabs: {wb.sheetnames}). Check the tab name/spelling."
        )

    records = _read_sheet_as_records(wb["Criteria"], ["Issue", "Assessment criteria"])
    if not records:
        return (
            f"Opened the 'Criteria' tab in {resolved}, but couldn't find a header row containing "
            f"both 'Issue' and 'Assessment criteria' within the first 15 rows, or there were no "
            f"data rows below it. Check the column headers match those exactly (case doesn't matter)."
        )
    return ""


def load_training_materials() -> dict:
    """Read the 'Training materials' tab and group examples by normalized issue name."""
    if not CRITERIA_XLSX.exists():
        return {}
    wb = openpyxl.load_workbook(CRITERIA_XLSX, data_only=True)
    if "Training materials" not in wb.sheetnames:
        return {}
    records = _read_sheet_as_records(
        wb["Training materials"],
        ["Photo", "Issue", "Defect or Good Practice", "Severity", "Location", "Observation"],
    )
    grouped: dict = {}
    for r in records:
        grouped.setdefault(_norm_issue(r["Issue"]), []).append(r)
    return grouped


def _build_checklist_block(criteria: list[dict], training_by_issue: dict) -> str:
    if not criteria:
        return "(No checklist found — library/QAQC_Agent_Training.xlsx is missing or its 'Criteria' tab is empty/unreadable.)"
    lines = []
    for i, c in enumerate(criteria, 1):
        issue = str(c.get("Issue") or "").strip()
        crit = str(c.get("Assessment criteria") or "").strip()
        lines.append(f"{i}. ISSUE: {issue}\n   Assessment criteria: {crit}")
        examples = training_by_issue.get(_norm_issue(issue), [])
        for ex in examples[:MAX_TRAINING_EXAMPLES_PER_ISSUE]:
            lines.append(
                f"   - Training example (Photo #{ex.get('Photo')}): "
                f"{ex.get('Defect or Good Practice')} (Severity: {ex.get('Severity')}) "
                f"at {ex.get('Location')} — {ex.get('Observation')}"
            )
    return "\n".join(lines)


# --------------------------------------------------------------------------------------
# Submittal context
# --------------------------------------------------------------------------------------

def _load_submittal_context(max_chars: int = MAX_SUBMITTAL_CHARS) -> str:
    if not SUBMITTALS_DIR.exists():
        return "(No /submittals folder found yet.)"
    parts = []
    for p in sorted(SUBMITTALS_DIR.glob("*")):
        if not p.is_file():
            continue
        if p.suffix.lower() in {".txt", ".md"}:
            try:
                text = p.read_text(errors="ignore")[:1200]
                parts.append(f"--- {p.name} ---\n{text}")
            except Exception:
                pass
        else:
            parts.append(f"[submittal on file, not text-extracted: {p.name} — cross-check manually if an assessment criterion needs it]")
    context = "\n\n".join(parts)
    return context[:max_chars] if context else "(No submittal files found in /submittals.)"


# --------------------------------------------------------------------------------------
# Stored feedback library (from past confirmed evaluations)
# --------------------------------------------------------------------------------------

def _ensure_library_workbook() -> None:
    """Create library/ and the feedback workbook with headers if they don't exist yet."""
    LIBRARY_DIR.mkdir(exist_ok=True)
    LIBRARY_PHOTOS_DIR.mkdir(exist_ok=True)
    if LIBRARY_XLSX.exists():
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Library"
    for i, h in enumerate(LIBRARY_HEADERS, 1):
        ws.cell(row=1, column=i, value=h)
    for i, w in enumerate([30, 26, 60, 14, 30, 14, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(LIBRARY_XLSX)


def _load_library_entries() -> list[dict]:
    if not LIBRARY_XLSX.exists():
        return []
    wb = openpyxl.load_workbook(LIBRARY_XLSX, data_only=True)
    ws = wb.active
    entries = []
    for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
        if not row or not row[0]:
            continue
        photo_filename, category, comment, severity, keywords, date_added, added_by = (
            list(row) + [None] * 7
        )[:7]
        entries.append({
            "photo_filename": photo_filename, "category": category, "comment": comment,
            "severity": severity, "keywords": keywords, "date_added": date_added, "added_by": added_by,
        })
    return entries


def _tokenize(text) -> set:
    return set(re.findall(r"[a-z0-9]+", str(text or "").lower()))


def _select_relevant_entries(entries: list[dict], query_text: str, max_entries: int = MAX_LIBRARY_MATCHES) -> list[dict]:
    if not entries:
        return []
    query_tokens = _tokenize(query_text)
    if not query_tokens:
        return sorted(entries, key=lambda e: str(e.get("date_added") or ""), reverse=True)[:max_entries]
    scored = []
    for e in entries:
        entry_tokens = _tokenize(e.get("category")) | _tokenize(e.get("comment")) | _tokenize(e.get("keywords"))
        scored.append((e, len(entry_tokens & query_tokens)))
    scored.sort(key=lambda pair: pair[1], reverse=True)
    top_scored = [e for e, score in scored if score > 0][:max_entries]
    if top_scored:
        return top_scored
    return sorted(entries, key=lambda e: str(e.get("date_added") or ""), reverse=True)[:max_entries]


def append_to_library(photo_path: Path, category: str, comment: str, severity: str = "",
                       keywords: str = "", added_by: str = "") -> Path:
    """Add a confirmed finding to the stored-feedback library so future evaluations reference it."""
    _ensure_library_workbook()
    dest_name = photo_path.name
    dest_path = LIBRARY_PHOTOS_DIR / dest_name
    if dest_path.exists() and dest_path.resolve() != photo_path.resolve():
        dest_name = f"{photo_path.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{photo_path.suffix}"
        dest_path = LIBRARY_PHOTOS_DIR / dest_name
    if dest_path.resolve() != photo_path.resolve():
        shutil.copy(photo_path, dest_path)

    wb = openpyxl.load_workbook(LIBRARY_XLSX)
    ws = wb.active
    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1, value=dest_name)
    ws.cell(row=next_row, column=2, value=category)
    ws.cell(row=next_row, column=3, value=comment)
    ws.cell(row=next_row, column=4, value=severity)
    ws.cell(row=next_row, column=5, value=keywords)
    ws.cell(row=next_row, column=6, value=date.today().isoformat())
    ws.cell(row=next_row, column=7, value=added_by)
    wb.save(LIBRARY_XLSX)
    return dest_path


# --------------------------------------------------------------------------------------
# Image handling
# --------------------------------------------------------------------------------------

def _encode_image(path: Path):
    mime, _ = mimetypes.guess_type(str(path))
    if mime is None:
        mime = "image/jpeg"
    data = base64.standard_b64encode(path.read_bytes()).decode("utf-8")
    return mime, data


# --------------------------------------------------------------------------------------
# Main evaluation
# --------------------------------------------------------------------------------------

def evaluate_photo(photo_path: Path, extra_context: str = "") -> str:
    """Evaluate one photo against the full Issue/Assessment Criteria checklist, grounded in
    training examples, stored feedback, and submittal context, with web search as a fallback
    for codes/standards and for catching issues outside the checklist."""
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise RuntimeError("ANTHROPIC_API_KEY is not set. Set it before running (see setup notes).")
    if api_key.strip() in {"sk-ant-...", "sk-ant-", ""}:
        raise RuntimeError(
            "ANTHROPIC_API_KEY is set to a placeholder, not a real key. "
            "Get a real key from console.anthropic.com and set it with setx (Windows)."
        )

    client = anthropic.Anthropic(api_key=api_key)
    mime, b64 = _encode_image(photo_path)

    criteria = load_criteria()
    training_by_issue = load_training_materials()
    checklist_block = _build_checklist_block(criteria, training_by_issue)

    submittal_context = _load_submittal_context()

    _ensure_library_workbook()
    library_entries = _load_library_entries()
    query_text = " ".join([extra_context or "", photo_path.stem.replace("_", " ").replace("-", " ")])
    feedback_matches = _select_relevant_entries(library_entries, query_text)
    if feedback_matches:
        fb_lines = [
            f"- [{e['category'] or 'Uncategorized'}] {e['comment']} "
            f"(Severity: {e['severity'] or 'n/a'}; Added: {e['date_added'] or 'n/a'})"
            for e in feedback_matches
        ]
        feedback_text = "STORED FEEDBACK from prior evaluations, matched to this photo/context:\n" + "\n".join(fb_lines)
    else:
        feedback_text = "No matching stored feedback found for this photo/context."

    content = [
        {"type": "text", "text": f"Photo to evaluate (filename: {photo_path.name}):"},
        {"type": "image", "source": {"type": "base64", "media_type": mime, "data": b64}},
    ]

    attached = 0
    for e in feedback_matches:
        if attached >= MAX_LIBRARY_IMAGES:
            break
        ref_name = e.get("photo_filename")
        if not ref_name:
            continue
        ref_path = LIBRARY_PHOTOS_DIR / str(ref_name)
        if ref_path.exists() and ref_path.is_file():
            try:
                ref_mime, ref_b64 = _encode_image(ref_path)
                content.append({"type": "text", "text": f"Stored feedback reference photo — matched finding: {e['comment']}"})
                content.append({"type": "image", "source": {"type": "base64", "media_type": ref_mime, "data": ref_b64}})
                attached += 1
            except Exception:
                pass

    user_text = (
        f"ISSUE CHECKLIST (evaluate every item, in order):\n{checklist_block}\n\n"
        f"{feedback_text}\n\n"
        f"SUBMITTAL CONTEXT:\n{submittal_context}\n"
    )
    if extra_context:
        user_text += f"\nAdditional field notes: {extra_context}\n"
    user_text += (
        "\nEvaluate ONLY the photo being evaluated (the first photo above) per your instructions. "
        "Any other photos are stored-feedback references for comparison — do not describe or "
        "evaluate them directly, only use them as grounding."
    )
    content.append({"type": "text", "text": user_text})

    response = client.messages.create(
        model=MODEL,
        max_tokens=3000,
        system=SYSTEM_PROMPT,
        tools=[{"type": "web_search_20250305", "name": "web_search"}],
        messages=[{"role": "user", "content": content}],
    )

    out = []
    for block in response.content:
        if block.type == "text":
            out.append(block.text)
    return "\n".join(out) if out else "(No text returned — check API response.)"


# --------------------------------------------------------------------------------------
# Report parsing — maps the structured report onto checklist columns
# --------------------------------------------------------------------------------------

_SECTION_PATTERNS = {
    "issues_found": re.compile(r"##\s*Issues Found\s*(.*?)(?=\n##|\Z)", re.DOTALL | re.IGNORECASE),
    "unclear": re.compile(r"##\s*Unclear\s*/\s*Field Team to Verify\s*(.*?)(?=\n##|\Z)", re.DOTALL | re.IGNORECASE),
    "no_issues": re.compile(r"##\s*No Issues Found\s*(.*?)(?=\n##|\Z)", re.DOTALL | re.IGNORECASE),
    "followup": re.compile(r"##\s*Follow-Up Recommended\s*(.*?)(?=\n##|\Z)", re.DOTALL | re.IGNORECASE),
}


def parse_report_sections(report_text: str) -> dict:
    """Pull the Issues Found / Unclear / No Issues Found / Follow-Up sections out of a report
    generated with the standard SYSTEM_PROMPT format. Empty string for any section not found."""
    sections = {"issues_found": "", "unclear": "", "no_issues": "", "followup": ""}
    for key, pattern in _SECTION_PATTERNS.items():
        m = pattern.search(report_text or "")
        if m:
            sections[key] = m.group(1).strip()
    return sections


def _section_has_content(text: str) -> bool:
    """True if a section actually contains something, vs. being empty or a bare 'None.'"""
    t = (text or "").strip().lower().rstrip(".")
    return bool(t) and t not in {"none", "none noted", "n/a", "none found"}


def notes_from_report(report_text: str) -> str:
    """Build the text for the checklist's Notes/Findings column: Issues Found, then Unclear,
    then a compact No Issues Found list, so the record is complete. Falls back to the full
    report if section parsing found nothing."""
    sections = parse_report_sections(report_text)
    parts = []
    if _section_has_content(sections["issues_found"]):
        parts.append(f"Issues Found:\n{sections['issues_found']}")
    if _section_has_content(sections["unclear"]):
        parts.append(f"Unclear / Field Team to Verify:\n{sections['unclear']}")
    if _section_has_content(sections["no_issues"]):
        parts.append(f"No Issues Found:\n{sections['no_issues']}")
    if parts:
        return "\n\n".join(parts)
    return report_text.strip()


def corrective_action_from_report(report_text: str) -> str:
    """Build the text for the checklist's Corrective Action Needed column: Follow-Up Recommended."""
    followup = parse_report_sections(report_text)["followup"]
    return followup if _section_has_content(followup) else ""


def suggest_status_from_report(report_text: str) -> str:
    """Suggest a checklist Status dropdown value from which sections actually have content.
    Always meant to be confirmed by the user, not written blind."""
    sections = parse_report_sections(report_text)
    if _section_has_content(sections["issues_found"]):
        return "Fail"
    if _section_has_content(sections["unclear"]):
        return "Needs Verification"
    return "Pass"
