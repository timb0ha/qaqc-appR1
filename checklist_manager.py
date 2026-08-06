"""
Generates the pre-made Construction QA/QC Checklist Template (Excel).
Run directly (python build_template.py) to (re)build templates/QAQC_Checklist_Template.xlsx,
or import build_template() to call it programmatically (used by app.py on first startup).
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
LEGEND_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
TITLE_FONT = Font(name="Arial", size=16, bold=True, color="1F4E78")
SUB_FONT = Font(name="Arial", size=10, italic=True, color="595959")
BODY_FONT = Font(name="Arial", size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def build_template(out_path: str = "templates/QAQC_Checklist_Template.xlsx") -> str:
    wb = openpyxl.Workbook()

    # ---------- Sheet 1: Instructions ----------
    ws0 = wb.active
    ws0.title = "Instructions"
    ws0["B2"] = "Construction QA/QC Checklist Template"
    ws0["B2"].font = TITLE_FONT
    ws0["B3"] = "Pre-made template — copy this file per inspection/area, then fill in the yellow cells."
    ws0["B3"].font = SUB_FONT

    instructions = [
        "1. Do not edit this master template directly — the app copies it to /checklists/ for you as a new, named checklist.",
        "2. Fill in the header block (Project, Location, Inspector, Date) on the 'Checklist' tab.",
        "3. For each line item, set Status via the dropdown: Pass / Fail / N/A / Needs Verification.",
        "4. Add notes and, where relevant, reference a photo filename uploaded through the app.",
        "5. Add or remove rows as needed — keep the same column structure so the app can read the file.",
        "6. Save. Saved checklists in /checklists/ are visible to all users of the shared folder.",
    ]
    r = 5
    for line in instructions:
        ws0.cell(row=r, column=2, value=line).font = BODY_FONT
        r += 1
    ws0.column_dimensions["A"].width = 3
    ws0.column_dimensions["B"].width = 110

    # ---------- Sheet 2: Checklist ----------
    ws = wb.create_sheet("Checklist")
    ws.sheet_view.showGridLines = False

    ws["B2"] = "CONSTRUCTION QA/QC INSPECTION CHECKLIST"
    ws["B2"].font = TITLE_FONT

    header_fields = [
        ("B4", "Project Name:"), ("B5", "Location / Area:"), ("B6", "Inspector:"),
        ("F4", "Date:"), ("F5", "Trade / CSI Division:"), ("F6", "Checklist ID:"),
    ]
    for cell, label in header_fields:
        col = cell[0]
        row = cell[1:]
        ws[cell] = label
        ws[cell].font = Font(name="Arial", size=10, bold=True)
        input_col = get_column_letter(openpyxl.utils.column_index_from_string(col) + 1)
        ws[f"{input_col}{row}"].fill = LEGEND_FILL
        ws[f"{input_col}{row}"].border = BORDER

    ws.merge_cells("B2:H2")

    table_start = 9
    cols = ["Item #", "CSI Division / Category", "Checklist Item / Requirement",
            "Status", "Photo Reference (filename)", "Notes / Findings", "Corrective Action Needed"]
    for i, c in enumerate(cols):
        cell = ws.cell(row=table_start, column=2 + i, value=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    sample_items = [
        ("1", "03 30 00 Concrete", "Formwork dimensions, bracing, and alignment verified against drawings", "", "", "", ""),
        ("2", "03 30 00 Concrete", "Rebar size, spacing, and cover verified prior to pour", "", "", "", ""),
        ("3", "04 20 00 Masonry", "Mortar joint tooling and consistency inspected", "", "", "", ""),
        ("4", "05 12 00 Structural Steel", "Bolted connections torqued and marked per spec", "", "", "", ""),
        ("5", "07 92 00 Joint Sealants", "Sealant application continuous, no voids or bubbling", "", "", "", ""),
        ("6", "09 90 00 Painting/Coating", "Surface prep and dry film thickness within spec range", "", "", "", ""),
    ]
    row = table_start + 1
    for item in sample_items:
        for j, val in enumerate(item):
            cell = ws.cell(row=row, column=2 + j, value=val)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        row += 1

    for _ in range(15):
        for j in range(len(cols)):
            cell = ws.cell(row=row, column=2 + j, value="")
            cell.border = BORDER
        row += 1

    last_row = row - 1
    dv = DataValidation(type="list", formula1='"Pass,Fail,N/A,Needs Verification"', allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(f"E{table_start+1}:E{last_row}")

    widths = {"A": 2, "B": 8, "C": 24, "D": 42, "E": 16, "F": 22, "G": 30, "H": 26}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = f"B{table_start+1}"

    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    import os
    os.makedirs("templates", exist_ok=True)
    print(f"Saved {build_template()}")
